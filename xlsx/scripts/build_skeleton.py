"""
Ewaantech XLSX Builder — clone-and-fill build skeleton.

Usage:
    1. Copy this file to your working dir as `build.py`
    2. Copy `lib.py` alongside it
    3. Place the white logo at `./logo_white.png` (pre-resize with lib.resize_logo)
    4. Set SRC and OUT paths below
    5. python3 build.py
    6. python3 /mnt/skills/public/xlsx/scripts/recalc.py <OUT>
    7. QA: convert to PDF, view every page

The skeleton handles the common workflow:
  - Inspect source workbook
  - Build a Cover sheet first
  - Rebuild data-table sheets (findings logs, trackers, etc.)
  - Rebuild non-tabular sheets (summary, manual, messages, etc.)
  - Apply consistent print setup

Modify the per-sheet build functions for your source's particulars.
"""

import os
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment

# Import the brand helpers
from lib import (
    RED, DARK_RED, NEAR_BLACK, CHARCOAL, GOLD, CREAM, PALE_CREAM, WHITE,
    fill, font, CENTER, CENTER_WRAP, LEFT_WRAP, LEFT_TOP, LEFT_CENTER,
    apply_title_bar, apply_header_row, apply_data_rows,
    apply_footer_cards, apply_end_of_findings_marker, apply_print_setup,
    default_column_width, split_findings_data, extract_footer_metadata,
    status_style, resize_logo,
    extract_source_images, embed_screenshot,
)

# =====================================================================
# CONFIGURE THESE
# =====================================================================

SRC = "/path/to/source.xlsx"      # CHANGE ME
OUT = "/mnt/user-data/outputs/Ewaantech_<DocumentType>_<Topic>.xlsx"  # CHANGE ME
LOGO_WHITE_SRC = "/path/to/EWT_LOGO_WHITE_2024.png"  # CHANGE ME

LOGO_WHITE_SMALL = "/home/claude/work/logo_white_small.png"

WORKBOOK_TITLE = "Bookly"
WORKBOOK_SUBTITLE = "Mobile App  ·  User Acceptance Testing Log"
WORKBOOK_EYEBROW = "UAT FINDINGS  /  2026"
WORKBOOK_DESCRIPTION = (
    "A consolidated record of findings, enhancements, and sign-offs across "
    "the UAT cycle. Use the sheet tabs below to navigate sprint logs.")

# =====================================================================
# HELPERS FOR THIS SKELETON
# =====================================================================

def read_sheet(src_wb, ws_name, max_rows=200, max_cols=60):
    """Walk the actual content of a sheet (ignores inflated max_row/max_col)."""
    ws = src_wb[ws_name]
    last_row, last_col = 0, 0
    for r in range(1, min(ws.max_row, max_rows) + 1):
        for c in range(1, min(ws.max_column, max_cols) + 1):
            v = ws.cell(row=r, column=c).value
            if v is not None and str(v).strip() != "":
                if r > last_row: last_row = r
                if c > last_col: last_col = c
    data = []
    for r in range(1, last_row + 1):
        row = []
        for c in range(1, last_col + 1):
            row.append(ws.cell(row=r, column=c).value)
        data.append(row)
    return data, last_row, last_col

# =====================================================================
# COVER SHEET BUILDER
# =====================================================================

def build_cover(wb, nav_entries, kpis):
    """Build the cover sheet.

    nav_entries: list of (sheet_name, description) tuples for the contents list.
    kpis: list of up to 4 (label, value, sublabel, dark?) tuples.
    """
    ws = wb.create_sheet("Cover")
    ws.sheet_view.showGridLines = False

    for col in range(1, 13):
        ws.column_dimensions[get_column_letter(col)].width = 12
    ws.row_dimensions[1].height = 18

    # Dark hero block (rows 2–12, cols B–K)
    for r in range(2, 13):
        for c in range(2, 12):
            ws.cell(row=r, column=c).fill = fill(NEAR_BLACK)

    # Red accent line (row 13, cols B–D)
    ws.row_dimensions[13].height = 4
    for c in range(2, 5):
        ws.cell(row=13, column=c).fill = fill(RED)

    # Hero content
    ws.row_dimensions[4].height = 22
    ws.merge_cells(start_row=4, start_column=3, end_row=4, end_column=10)
    cc = ws.cell(row=4, column=3, value=WORKBOOK_EYEBROW)
    cc.font = font(size=11, bold=True, color=GOLD)
    cc.alignment = Alignment(horizontal="left", vertical="center")
    cc.fill = fill(NEAR_BLACK)

    ws.row_dimensions[5].height = 12
    ws.row_dimensions[6].height = 56
    ws.row_dimensions[7].height = 56

    ws.merge_cells(start_row=6, start_column=3, end_row=6, end_column=10)
    cc = ws.cell(row=6, column=3, value=WORKBOOK_TITLE)
    cc.font = font(size=48, bold=True, color=WHITE)
    cc.alignment = Alignment(horizontal="left", vertical="center")
    cc.fill = fill(NEAR_BLACK)

    ws.merge_cells(start_row=7, start_column=3, end_row=7, end_column=10)
    cc = ws.cell(row=7, column=3, value=WORKBOOK_SUBTITLE)
    cc.font = font(size=16, color=CREAM)
    cc.alignment = Alignment(horizontal="left", vertical="center")
    cc.fill = fill(NEAR_BLACK)

    ws.row_dimensions[9].height = 12
    ws.row_dimensions[10].height = 22
    ws.row_dimensions[11].height = 22
    ws.merge_cells(start_row=10, start_column=3, end_row=11, end_column=10)
    cc = ws.cell(row=10, column=3, value=WORKBOOK_DESCRIPTION)
    cc.font = font(size=11, color=CREAM)
    cc.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    cc.fill = fill(NEAR_BLACK)

    if os.path.exists(LOGO_WHITE_SMALL):
        try:
            img = XLImage(LOGO_WHITE_SMALL)
            ws.add_image(img, "I3")
        except Exception:
            pass

    # KPI tile strip (rows 14–18)
    ws.row_dimensions[14].height = 14
    ws.row_dimensions[15].height = 20
    ws.row_dimensions[16].height = 56
    ws.row_dimensions[17].height = 22
    ws.row_dimensions[18].height = 6

    col_pairs = [(2, 3), (4, 5), (6, 7), (8, 9)]
    while len(kpis) < 4:
        kpis = list(kpis) + [(None, None, None, None)]
    for i, ((c1, c2), spec) in enumerate(zip(col_pairs, kpis)):
        label, num, sub, dark = spec
        if label is None:
            continue
        bg = NEAR_BLACK if dark else CREAM
        fg = WHITE if dark else NEAR_BLACK
        eb_col = GOLD if dark else RED
        sub_col = CREAM if dark else CHARCOAL

        for k in range(c1, c2 + 1):
            ws.cell(row=15, column=k).fill = fill(bg)
            ws.cell(row=16, column=k).fill = fill(bg)
            ws.cell(row=17, column=k).fill = fill(bg)
            ws.cell(row=18, column=k).fill = fill(RED)

        ws.merge_cells(start_row=15, start_column=c1, end_row=15, end_column=c2)
        cc = ws.cell(row=15, column=c1, value=label.upper())
        cc.font = font(size=9, bold=True, color=eb_col)
        cc.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        ws.merge_cells(start_row=16, start_column=c1, end_row=16, end_column=c2)
        cc = ws.cell(row=16, column=c1, value=str(num))
        cc.font = font(size=44, bold=True, color=fg)
        cc.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        ws.merge_cells(start_row=17, start_column=c1, end_row=17, end_column=c2)
        cc = ws.cell(row=17, column=c1, value=sub or "")
        cc.font = font(size=9, color=sub_col)
        cc.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    # Contents nav
    ws.row_dimensions[19].height = 24
    ws.row_dimensions[20].height = 22
    cc = ws.cell(row=20, column=2, value="CONTENTS")
    cc.font = font(size=10, bold=True, color=RED)
    cc.alignment = Alignment(horizontal="left", vertical="center")

    ws.row_dimensions[21].height = 3
    for c in range(2, 4):
        ws.cell(row=21, column=c).fill = fill(RED)

    for i, (sheet_name, desc) in enumerate(nav_entries):
        r = 22 + i
        ws.row_dimensions[r].height = 24
        bg = WHITE if i % 2 == 0 else PALE_CREAM
        for c in range(2, 11):
            ws.cell(row=r, column=c).fill = fill(bg)

        cc = ws.cell(row=r, column=2, value="›")
        cc.font = font(size=14, bold=True, color=RED)
        cc.alignment = Alignment(horizontal="center", vertical="center")

        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
        cc = ws.cell(row=r, column=3, value=sheet_name)
        cc.hyperlink = f"#'{sheet_name}'!A1"
        cc.style = "Hyperlink"
        cc.font = font(size=11, bold=True, color=NEAR_BLACK)
        cc.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=10)
        cc = ws.cell(row=r, column=6, value=desc)
        cc.font = font(size=10, color=CHARCOAL)
        cc.alignment = Alignment(horizontal="left", vertical="center")

    apply_print_setup(ws, orientation="P", fit_to_height=1, repeat_title=None)
    return ws

# =====================================================================
# DATA-TABLE SHEET BUILDER  (findings, trackers, issue logs)
# =====================================================================

def build_data_table_sheet(wb, sheet_name, eyebrow, title, headers,
                           data_rows, footer_rows=None,
                           status_col_name="Status", sl_no_col_name="Sl No"):
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False

    n_cols = len(headers)
    for i, h in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(i)].width = default_column_width(h)

    apply_title_bar(ws, eyebrow, title, n_cols, logo_path=LOGO_WHITE_SMALL)
    apply_header_row(ws, headers, row_num=4)

    # Resolve special column indices
    try:
        status_idx = headers.index(status_col_name) + 1
    except ValueError:
        status_idx = None
    try:
        sl_idx = headers.index(sl_no_col_name) + 1
    except ValueError:
        sl_idx = None
        if "" in headers and sl_no_col_name == "":
            sl_idx = headers.index("") + 1

    date_col_indices = set()
    for i, h in enumerate(headers, start=1):
        hl = (h or "").lower()
        if hl in ("date added", "due date", "enhancement task(s) approval"):
            date_col_indices.add(i)

    apply_data_rows(ws, data_rows, start_row=5, n_cols=n_cols,
                    status_col_idx=status_idx, sl_no_col_idx=sl_idx,
                    date_col_indices=date_col_indices)

    last_data_row = 4 + len(data_rows)
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:{get_column_letter(n_cols)}{last_data_row}"

    end_row = last_data_row + 2
    if footer_rows:
        ws.row_dimensions[last_data_row + 1].height = 14
        block_start = last_data_row + 2
        meta = extract_footer_metadata(footer_rows)
        apply_footer_cards(ws, block_start, n_cols, **meta)
        end_row = block_start + 5

    apply_end_of_findings_marker(ws, end_row, n_cols)
    apply_print_setup(ws, orientation="L", fit_to_height=0,
                      repeat_title="1:4")
    return ws

# =====================================================================
# EXAMPLE MAIN  (adapt to your source)
# =====================================================================

def main():
    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    os.makedirs("/home/claude/work", exist_ok=True)

    # Pre-resize the logo
    resize_logo(LOGO_WHITE_SRC, LOGO_WHITE_SMALL, target_h_px=36)

    src_wb = load_workbook(SRC, data_only=True)
    wb = Workbook()
    wb.remove(wb.active)

    # === Cover ===
    nav = [(name, "...description...") for name in src_wb.sheetnames]
    kpis = [
        ("Total Hours",       "108",  "Logged across all sprints", True),
        ("Completed",         "27.5", "Hours on closed tasks",      False),
        ("Awaiting Approval", "80.5", "Hours pending sign-off",     False),
        ("Sprints",           "5",    "UAT cycles tracked",         True),
    ]
    build_cover(wb, nav, kpis)

    # === Per source-sheet builders ===
    for src_name in src_wb.sheetnames:
        rows, _, _ = read_sheet(src_wb, src_name)
        if not rows:
            continue
        headers = ["" if v is None else str(v) for v in rows[0]]
        body = rows[1:]
        data, footer = split_findings_data(body)
        build_data_table_sheet(
            wb, src_name,
            eyebrow="UAT Sprint",        # customize per sheet
            title=src_name,              # customize per sheet
            headers=headers,
            data_rows=data,
            footer_rows=footer,
        )

    # === Re-embed source screenshots ===
    # openpyxl doesn't surface embedded images; we recover them ourselves.
    placements = extract_source_images(SRC, extract_dir="/home/claude/work/imgs")
    for sheet_name, items in placements.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for it in items:
            # Source row R → rebuild row (R + 3): branded title bar occupies
            # 3 rows above the header at row 4.
            embed_screenshot(
                ws, image_path=it["image_path"],
                target_row=it["source_row"] + 3,
                target_col_header="Screenshot",
            )

    # Tab colors
    for ws in wb.worksheets:
        ws.sheet_properties.tabColor = RED if ws.title in ("Cover", "Summary") else CHARCOAL

    wb.save(OUT)
    print(f"Saved: {OUT}")

if __name__ == "__main__":
    main()
