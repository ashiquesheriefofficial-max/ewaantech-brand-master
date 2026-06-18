"""
Ewaantech XLSX Builder — reusable openpyxl style helpers.

This library is the canonical scaffold for every Ewaantech-branded .xlsx.
Import from it in your build script — do NOT redefine brand tokens or
helpers elsewhere. If the brand evolves, update this file only.

Validated against the approved UAT Findings Bookly build.
"""

import os
import datetime as _dt
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

try:
    from PIL import Image as PILImage
except ImportError:
    PILImage = None

# =====================================================================
# BRAND TOKENS  (from Corporate_Guideline_2026.pdf — verify each session)
# =====================================================================

RED         = "CE202F"
DARK_RED    = "AC1F24"
NEAR_BLACK  = "221E1F"
CHARCOAL    = "242424"
GOLD        = "C1A068"
CREAM       = "EFEBE3"
PALE_CREAM  = "F8F5EE"    # alt-row zebra color, half-strength cream
SUBTLE_GREY = "DDD8CF"    # 1px row-border color
WHITE       = "FFFFFF"

BRAND_FONT  = "Calibri"   # universal fallback for system rendering

# =====================================================================
# STYLE PRIMITIVES
# =====================================================================

def fill(hex_):
    """Solid-color fill from a 6-char HEX (no '#')."""
    return PatternFill("solid", fgColor=hex_)

def font(size=10, bold=False, color=NEAR_BLACK, name=BRAND_FONT, italic=False):
    """Calibri-default font with brand-safe color."""
    return Font(name=name, size=size, bold=bold, italic=italic, color=color)

# Alignments
CENTER       = Alignment(horizontal="center", vertical="center")
CENTER_WRAP  = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_CENTER  = Alignment(horizontal="left",   vertical="center", indent=1)
LEFT_WRAP    = Alignment(horizontal="left",   vertical="center", wrap_text=True, indent=1)
LEFT_TOP     = Alignment(horizontal="left",   vertical="top",    wrap_text=True, indent=1)

# Borders
NO_BORDER = Border()
ROW_BOTTOM_BORDER = Border(
    bottom=Side(style="thin", color=SUBTLE_GREY),
    left=Side(style="thin", color="F0ECE4"),
    right=Side(style="thin", color="F0ECE4"),
)
HEADER_BORDER = Border(
    left=Side(style="thin", color=DARK_RED),
    right=Side(style="thin", color=DARK_RED),
    top=Side(style="medium", color=RED),
    bottom=Side(style="medium", color=DARK_RED),
)

# =====================================================================
# LOGO HELPERS
# =====================================================================

def resize_logo(src_path, dst_path, target_h_px=36):
    """Pre-resize the logo to a small embedded size.

    Excel embeds the full-resolution image otherwise — bloats the file and
    renders blurry. Returns (width, height) of the resized image.

    Requires Pillow. If Pillow is unavailable, returns None and the caller
    should embed the source as-is (suboptimal but functional).
    """
    if PILImage is None:
        return None
    img = PILImage.open(src_path)
    w, h = img.size
    ratio = target_h_px / h
    new_w = int(w * ratio)
    img = img.resize((new_w, target_h_px), PILImage.LANCZOS)
    img.save(dst_path, "PNG", optimize=True)
    return new_w, target_h_px

# =====================================================================
# TITLE BAR  (the signature motif of every Ewaantech sheet)
# =====================================================================

def apply_title_bar(ws, eyebrow, title, n_cols, logo_path=None):
    """Apply the standard branded title bar to a worksheet.

    Layout:
      Row 1 (h=44) — dark near-black, "EYEBROW · Title" in white at left,
                     white logo embedded at top-right
      Row 2 (h=4)  — red accent strip
      Row 3 (h=8)  — spacer

    `n_cols` is the number of columns the bar should span (the sheet's
    data column count). The eyebrow is upper-cased automatically.
    """
    if n_cols < 4:
        n_cols = 4

    ws.row_dimensions[1].height = 44
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    c = ws.cell(row=1, column=1)
    c.fill = fill(NEAR_BLACK)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    c.value = f"  {eyebrow.upper()}   ·   {title}"
    c.font = font(size=14, bold=True, color=WHITE)
    for cc in range(2, n_cols + 1):
        ws.cell(row=1, column=cc).fill = fill(NEAR_BLACK)

    ws.row_dimensions[2].height = 4
    for cc in range(1, n_cols + 1):
        ws.cell(row=2, column=cc).fill = fill(RED)

    ws.row_dimensions[3].height = 8

    if logo_path and os.path.exists(logo_path):
        try:
            img = XLImage(logo_path)
            anchor_col = max(1, n_cols - 1)
            ws.add_image(img, f"{get_column_letter(anchor_col)}1")
        except Exception:
            pass

# =====================================================================
# HEADER ROW
# =====================================================================

def apply_header_row(ws, headers, row_num=4):
    """Brand-red header row, white bold center-wrap text, h=38."""
    ws.row_dimensions[row_num].height = 38
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row_num, column=i, value=h)
        c.fill = fill(RED)
        c.font = font(size=10, bold=True, color=WHITE)
        c.alignment = CENTER_WRAP
        c.border = HEADER_BORDER

# =====================================================================
# STATUS PILL MAPPING  (palette-locked — never invent a new color)
# =====================================================================

def status_style(status):
    """Return (fill_hex, font_hex, bold) for a status string.

    All colors are from the brand palette only. Returns None for empty
    values so the cell stays in the default zebra-stripe state.
    """
    if status is None:
        return None
    s = str(status).strip().lower()
    if not s:
        return None
    if s == "closed":
        return (CHARCOAL, CREAM, True)
    if s == "done":
        return (GOLD, NEAR_BLACK, True)
    if s in ("open", "re-open", "reopen"):
        return (RED, WHITE, True)
    if s in ("in-progress", "in progress"):
        return (DARK_RED, WHITE, True)
    if s == "review":
        return (NEAR_BLACK, GOLD, True)
    if s in ("on-hold", "on hold", "not an issue"):
        return (CREAM, CHARCOAL, True)
    return (CREAM, CHARCOAL, False)

# =====================================================================
# DATA ROWS
# =====================================================================

def apply_data_rows(ws, data_rows, start_row, n_cols,
                    status_col_idx=None, sl_no_col_idx=None,
                    date_col_indices=None, row_height=38):
    """Write data rows with alternating cream/white fills and status pills.

    `data_rows` — list of lists; missing cells handled gracefully.
    `start_row` — first row of data (typically 5, since rows 1–4 are title+header).
    `status_col_idx` — 1-based column index for the Status column (gets pill styling).
    `sl_no_col_idx` — 1-based column index for the Sl-No column (centered, bold).
    `date_col_indices` — set of 1-based column indices to format as dates.
    """
    date_col_indices = date_col_indices or set()

    for i, row in enumerate(data_rows):
        r = start_row + i
        bg = WHITE if i % 2 == 0 else PALE_CREAM
        ws.row_dimensions[r].height = row_height

        for j in range(1, n_cols + 1):
            val = row[j-1] if j-1 < len(row) else None
            c = ws.cell(row=r, column=j, value=val)
            c.fill = fill(bg)
            c.font = font(size=10, color=NEAR_BLACK)

            if sl_no_col_idx and j == sl_no_col_idx:
                c.alignment = CENTER
                c.font = font(size=10, bold=True, color=NEAR_BLACK)
            else:
                c.alignment = LEFT_TOP
            c.border = ROW_BOTTOM_BORDER

            # Date formatting
            is_date = (isinstance(val, (_dt.datetime, _dt.date)) or
                       j in date_col_indices)
            if is_date:
                c.number_format = "dd mmm yyyy"
                if j != (status_col_idx or -1):
                    c.alignment = CENTER

            # Status pill (must run last to override)
            if status_col_idx and j == status_col_idx:
                st = status_style(val)
                if st is not None:
                    fhex, txt_color, bold = st
                    c.fill = fill(fhex)
                    c.font = font(size=10, bold=bold, color=txt_color)
                    c.alignment = CENTER

# =====================================================================
# FOOTER CARDS  (Total Hours + Sprint Details)
# =====================================================================

def apply_footer_cards(ws, start_row, n_cols, total_hours=None,
                       sprint_name=None, submitted_end=None,
                       sprint_start=None, sprint_end=None):
    """Draw the dark total-hours card (left) and cream sprint-details card (right).

    Returns the row index of the red strip beneath the cards.
    """
    # Total-hours card  (cols 1–3)
    ws.row_dimensions[start_row].height = 18      # eyebrow
    ws.row_dimensions[start_row + 1].height = 50  # numeral
    ws.row_dimensions[start_row + 2].height = 18  # subtitle
    ws.row_dimensions[start_row + 3].height = 4   # red strip

    for r in range(start_row, start_row + 3):
        for c in range(1, 4):
            ws.cell(row=r, column=c).fill = fill(NEAR_BLACK)

    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=3)
    cc = ws.cell(row=start_row, column=1, value="  TOTAL HOURS")
    cc.font = font(size=9, bold=True, color=GOLD)
    cc.alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells(start_row=start_row+1, start_column=1, end_row=start_row+1, end_column=3)
    num_str = (f"{total_hours:g}" if isinstance(total_hours, (int, float)) else "—")
    cc = ws.cell(row=start_row+1, column=1, value=f"  {num_str}")
    cc.font = font(size=36, bold=True, color=WHITE)
    cc.alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells(start_row=start_row+2, start_column=1, end_row=start_row+2, end_column=3)
    cc = ws.cell(row=start_row+2, column=1, value="  Estimated effort for this sprint")
    cc.font = font(size=9, color=CREAM)
    cc.alignment = Alignment(horizontal="left", vertical="center")

    for c in range(1, 4):
        ws.cell(row=start_row+3, column=c).fill = fill(RED)

    # Sprint-details card  (cols 5 onward)
    if n_cols >= 6:
        info_c1 = 5
        info_c2 = min(n_cols, 11)
        for r in range(start_row, start_row + 3):
            for c in range(info_c1, info_c2 + 1):
                ws.cell(row=r, column=c).fill = fill(CREAM)
        ws.merge_cells(start_row=start_row, start_column=info_c1,
                       end_row=start_row, end_column=info_c2)
        cc = ws.cell(row=start_row, column=info_c1, value="  SPRINT DETAILS")
        cc.font = font(size=9, bold=True, color=RED)
        cc.alignment = Alignment(horizontal="left", vertical="center")

        ws.merge_cells(start_row=start_row+1, start_column=info_c1,
                       end_row=start_row+2, end_column=info_c2)
        info_bits = []
        if sprint_name:
            info_bits.append(str(sprint_name))
        if submitted_end:
            info_bits.append(f"Submitted End Date: " +
                (submitted_end.strftime('%d %b %Y') if hasattr(submitted_end, 'strftime') else str(submitted_end)))
        if sprint_start:
            info_bits.append(f"Sprint Start: " +
                (sprint_start.strftime('%d %b %Y') if hasattr(sprint_start, 'strftime') else str(sprint_start)))
        if sprint_end:
            info_bits.append(f"Sprint End: " +
                (sprint_end.strftime('%d %b %Y') if hasattr(sprint_end, 'strftime') else str(sprint_end)))
        info_text = "\n".join(info_bits)
        cc = ws.cell(row=start_row+1, column=info_c1, value=info_text)
        cc.font = font(size=10, color=NEAR_BLACK)
        cc.alignment = Alignment(horizontal="left", vertical="center",
                                 wrap_text=True, indent=1)
        for c in range(info_c1, info_c2 + 1):
            ws.cell(row=start_row+3, column=c).fill = fill(CREAM)

    return start_row + 3

def apply_end_of_findings_marker(ws, row, n_cols):
    """Pale-cream centered marker — italic ‘— END OF FINDINGS LIST —’."""
    ws.row_dimensions[row].height = 22
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
    cc = ws.cell(row=row, column=1, value="—  END OF FINDINGS LIST  —")
    cc.font = font(size=10, bold=True, color=CHARCOAL, italic=True)
    cc.alignment = Alignment(horizontal="center", vertical="center")
    for c in range(1, n_cols + 1):
        ws.cell(row=row, column=c).fill = fill(PALE_CREAM)

# =====================================================================
# PAGE SETUP
# =====================================================================

def apply_print_setup(ws, orientation="L", fit_to_height=0, repeat_title="1:4"):
    """Apply landscape A4, fit-to-width, repeat the title-bar + header on print.

    orientation — "L" landscape (data tables), "P" portrait (editorial).
    fit_to_height — 0 to allow vertical paging, 1 to force whole sheet on one page.
    repeat_title — print_title_rows range, default rows 1–4 (the brand title bar
                   + header).
    """
    ws.page_setup.orientation = (ws.ORIENTATION_LANDSCAPE if orientation == "L"
                                 else ws.ORIENTATION_PORTRAIT)
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = fit_to_height
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.3
    ws.page_margins.right = 0.3
    ws.page_margins.top = 0.4
    ws.page_margins.bottom = 0.4
    ws.page_margins.header = 0.2
    ws.page_margins.footer = 0.2
    if repeat_title:
        ws.print_title_rows = repeat_title
    ws.sheet_view.showGridLines = False

# =====================================================================
# COLUMN WIDTHS  (from the approved Bookly build)
# =====================================================================

def default_column_width(header_text):
    """Return the brand-standard width for a column based on its header name.

    Headers not in this map should be sized to fit their content (12–22 is
    a safe default for unknown text columns).
    """
    if header_text is None:
        return 7   # blank header → assume Sl-No column
    hl = str(header_text).lower().strip()
    if hl in ("", "sl no"):           return 7
    if hl == "milestone":             return 14
    if hl == "section":               return 20
    if hl == "findings":              return 46
    if hl == "screenshot":            return 18
    if hl == "client comment":        return 22
    if hl == "date added":            return 15
    if hl in ("finding type",):       return 18
    if hl == "enhancement task(s) approval":  return 22
    if hl == "sprint name":           return 16
    if "estimated hours" in hl:       return 15
    if hl == "status":                return 14
    if hl == "due date":              return 15
    if "ewaantech comment" in hl:     return 26
    if "qc comment" in hl:            return 22
    if "dev comment" in hl:           return 22
    return 15

# =====================================================================
# SPLITTING SOURCE DATA INTO DATA + FOOTER
# =====================================================================

def split_findings_data(rows):
    """Split a body of rows (excluding the header row) into (data, footer).

    Footer starts at the first row containing markers like "END OF FINDINGS",
    "Submitted End Date", "Sprint Name:", etc. Also handles the case where a
    lone numeric value (the total-hours) sits in a row at the end of data
    without any 'total' label — that row gets moved to the footer.
    """
    data, footer = [], []
    found_footer = False
    for row in rows:
        joined = " | ".join("" if v is None else str(v) for v in row).lower()
        if (("end of findings" in joined) or ("submitted end date" in joined)
                or ("sprint name:" in joined) or ("sprint start date" in joined)
                or ("sprint end date" in joined)):
            found_footer = True
        if found_footer:
            footer.append(row)
        else:
            if any(v is not None and str(v).strip() != "" for v in row):
                data.append(row)
            elif data:
                data.append(row)

    while data and not any(v is not None and str(v).strip() != "" for v in data[-1]):
        data.pop()

    if data:
        last = data[-1]
        non_empty = [v for v in last if v is not None and str(v).strip() != ""]
        if (len(non_empty) == 1 and isinstance(non_empty[0], (int, float))):
            footer.insert(0, last)
            data.pop()
    return data, footer

def extract_footer_metadata(footer_rows):
    """Walk footer rows for the total-hours, sprint name, and date markers.

    Returns a dict with keys: total_hours, sprint_name, submitted_end,
    sprint_start, sprint_end. Missing values are None.
    """
    meta = dict(total_hours=None, sprint_name=None,
                submitted_end=None, sprint_start=None, sprint_end=None)
    for fr in footer_rows:
        joined = " | ".join("" if v is None else str(v) for v in fr).lower()

        # any numeric value in a footer row is a candidate total
        for v in fr:
            if isinstance(v, (int, float)) and meta["total_hours"] is None:
                meta["total_hours"] = v
                break

        if "submitted end date" in joined:
            for v in fr:
                if hasattr(v, 'strftime'):
                    meta["submitted_end"] = v
                    break

        if "sprint name" in joined:
            for v in fr:
                if isinstance(v, str):
                    s = v.strip().lower()
                    if "sprint name" in s or s.endswith(":"):
                        continue
                    if "sprint" in s:
                        meta["sprint_name"] = v.strip()
                        break

        if "sprint start" in joined:
            for v in fr:
                if hasattr(v, 'strftime'):
                    meta["sprint_start"] = v
                    break

        if "sprint end" in joined:
            for v in fr:
                if hasattr(v, 'strftime'):
                    meta["sprint_end"] = v
                    break
    return meta

# =====================================================================
# EMBEDDED IMAGE HANDLING
# (openpyxl's public API does NOT surface images on data_only=True loads;
# we read the zip ourselves to recover them.)
# =====================================================================

def extract_source_images(src_xlsx, extract_dir):
    """Parse a source .xlsx as a zip; recover every embedded image with anchor.

    Returns a dict keyed by sheet name:
      {
        "SheetName": [
          {"source_row": int (1-based),
           "source_col": int (1-based),
           "image": "image1.jpeg",
           "image_path": "/extract/dir/image1.jpeg"},
          ...
        ],
        ...
      }

    Sheets without images simply aren't in the dict. Returns {} if the
    workbook has no embedded media at all.
    """
    import zipfile
    import xml.etree.ElementTree as ET

    os.makedirs(extract_dir, exist_ok=True)
    placements = {}

    with zipfile.ZipFile(src_xlsx) as z:
        names = set(z.namelist())

        # workbook.xml → sheet list with relationship ids
        ns_main = '{http://schemas.openxmlformats.org/spreadsheetml/2006/main}'
        rid_ns = '{http://schemas.openxmlformats.org/officeDocument/2006/relationships}'
        wb_xml = ET.fromstring(z.read('xl/workbook.xml').decode())
        sheets = [(s.get('name'), s.get(rid_ns + 'id'))
                  for s in wb_xml.findall(f'{ns_main}sheets/{ns_main}sheet')]

        # workbook rels → sheet file paths
        wb_rels = ET.fromstring(z.read('xl/_rels/workbook.xml.rels').decode())
        rid_to_path = {r.get('Id'): r.get('Target') for r in wb_rels}
        sheet_to_file = {n: rid_to_path[rid] for n, rid in sheets if rid in rid_to_path}

        ns_xdr = {'xdr': 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing',
                  'a':   'http://schemas.openxmlformats.org/drawingml/2006/main'}

        for sheet_name, sheet_file in sheet_to_file.items():
            base = sheet_file.split('/')[-1]            # e.g. sheet1.xml
            rels_path = f'xl/worksheets/_rels/{base}.rels'
            if rels_path not in names:
                continue
            rels = ET.fromstring(z.read(rels_path).decode())
            drawing_target = None
            for r in rels:
                if 'drawing' in r.get('Target', ''):
                    drawing_target = r.get('Target')
                    break
            if not drawing_target:
                continue
            drawing_file = os.path.basename(drawing_target)
            drawing_xml = ET.fromstring(z.read(f'xl/drawings/{drawing_file}').decode())
            drawing_rels = ET.fromstring(
                z.read(f'xl/drawings/_rels/{drawing_file}.rels').decode())
            rid_to_media = {r.get('Id'): os.path.basename(r.get('Target'))
                            for r in drawing_rels}

            items = []
            for anchor in drawing_xml:
                frm = anchor.find('xdr:from', ns_xdr)
                if frm is None:
                    continue
                col = int(frm.find('xdr:col', ns_xdr).text)   # 0-based
                row = int(frm.find('xdr:row', ns_xdr).text)
                blip = anchor.find('.//a:blip', ns_xdr)
                rid = blip.get(rid_ns + 'embed') if blip is not None else None
                media = rid_to_media.get(rid)
                if not media:
                    continue
                items.append({
                    'source_row': row + 1,
                    'source_col': col + 1,
                    'image': media,
                    'image_path': os.path.join(extract_dir, media),
                })
            if items:
                placements[sheet_name] = items

        # Dump every media file to disk
        for n in z.namelist():
            if n.startswith('xl/media/'):
                out = os.path.join(extract_dir, os.path.basename(n))
                with open(out, 'wb') as f:
                    f.write(z.read(n))

    return placements


def _resize_image_to_box(src_path, dst_path, max_w_px, max_h_px):
    """Resize keeping aspect ratio to fit a max bounding box. Returns (w, h)."""
    if PILImage is None:
        return None
    img = PILImage.open(src_path)
    if img.mode in ('RGBA', 'P'):
        img = img.convert('RGB')
    ow, oh = img.size
    scale = min(max_w_px / ow, max_h_px / oh)
    nw = max(1, int(ow * scale))
    nh = max(1, int(oh * scale))
    img = img.resize((nw, nh), PILImage.LANCZOS)
    img.save(dst_path, "JPEG", quality=85, optimize=True)
    return nw, nh


def embed_screenshot(ws, image_path, target_row, target_col_header="Screenshot",
                     max_w_px=140, max_h_px=130, column_width=22):
    """Embed an image into a specific cell of a data-table sheet.

    - Resolves `target_col_header` against the header row at row 4.
      Falls back to column 5 (the standard Screenshot position) if not found.
    - Pre-resizes the image to fit within max_w_px × max_h_px (PIL).
    - Widens the column to `column_width` and grows the target row to fit
      the image plus 6 px padding.
    - Embeds with a one-cell anchor so the image stays clipped to the cell.

    Other rows are not affected, so non-image rows stay at the standard
    38-pt height — only image-bearing rows grow.
    """
    from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
    from openpyxl.utils.units import pixels_to_EMU
    from openpyxl.drawing.xdr import XDRPositiveSize2D

    # Resolve column from header row
    target_col = None
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=4, column=c).value
        if isinstance(v, str) and v.strip().lower() == target_col_header.lower():
            target_col = c
            break
    if target_col is None:
        target_col = 5

    ws.column_dimensions[get_column_letter(target_col)].width = column_width

    # Resize the image to a temp file alongside the source
    dst = image_path + ".resized.jpg"
    dims = _resize_image_to_box(image_path, dst, max_w_px, max_h_px)
    if dims is None:
        return
    w_px, h_px = dims

    # Grow row height (pt) to fit the image + small padding
    min_h_pt = (h_px + 6) * 0.75
    cur_h = ws.row_dimensions[target_row].height or 38
    if cur_h < min_h_pt:
        ws.row_dimensions[target_row].height = min_h_pt

    # Anchor inside the cell with a small left/top inset
    img = XLImage(dst)
    img.width = w_px
    img.height = h_px
    anchor = OneCellAnchor()
    anchor._from = AnchorMarker(
        col=target_col - 1, colOff=pixels_to_EMU(4),
        row=target_row - 1, rowOff=pixels_to_EMU(3),
    )
    anchor.ext = XDRPositiveSize2D(cx=pixels_to_EMU(w_px), cy=pixels_to_EMU(h_px))
    img.anchor = anchor
    ws.add_image(img)
